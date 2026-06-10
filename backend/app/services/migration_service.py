"""换机助手 — 数据迁移服务

支持从其他代账软件导出 Excel 批量导入，也支持本产品完整数据导出/导入。
"""
import uuid
import json
import io
from datetime import datetime, date
from typing import Any
from sqlalchemy.orm import Session
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.models.client import Client
from app.models.account import ChartOfAccount
from app.models.voucher import AccountingVoucher
from app.models.invoice import Invoice
from app.models.filing import TaxFiling
from app.models.contract import Contract
from app.models.fixed_asset import FixedAsset
from app.models.user import User, UserClientAssignment
from app.models.subscription import ClientSubscription
from app.models.document import OriginalDocument

# ============================================================
# 模板定义
# ============================================================

SHEET_DEFS = {
    "客户信息": {
        "model": Client,
        "columns": ["name", "tax_no", "taxpayer_type", "industry", "address",
                     "contact_person", "contact_phone", "service_start", "service_end", "remark", "is_active"],
        "required": ["name", "tax_no"],
        "unique": ["tax_no"],
    },
    "会计科目": {
        "model": ChartOfAccount,
        "columns": ["code", "name", "category", "parent_code", "direction", "is_active"],
        "required": ["code", "name", "category"],
        "unique": ["code"],
    },
    "记账凭证": {
        "model": AccountingVoucher,
        "columns": ["voucher_no", "voucher_date", "summary", "entries",
                     "total_debit", "total_credit", "status", "client_id", "created_by"],
        "required": ["voucher_no", "voucher_date", "summary", "entries"],
        "unique": ["voucher_no"],
    },
    "发票": {
        "model": Invoice,
        "columns": ["client_id", "buyer_name", "buyer_tax_no", "buyer_address", "buyer_phone",
                     "buyer_bank", "buyer_account", "invoice_type", "items",
                     "total_amount", "total_tax", "grand_total", "remark", "status",
                     "invoice_code", "invoice_no", "created_by"],
        "required": ["client_id", "buyer_name", "buyer_tax_no", "items"],
        "unique": [],
    },
    "申报记录": {
        "model": TaxFiling,
        "columns": ["tax_type", "period", "client_id", "filing_result",
                     "status", "filed_at"],
        "required": ["tax_type", "period", "client_id"],
        "unique": [],
    },
    "合同": {
        "model": Contract,
        "columns": ["client_id", "contract_no", "contract_name", "contract_type",
                     "counterparty", "amount", "start_date", "end_date", "status",
                     "payment_terms", "revenue_period", "monthly_revenue", "remark"],
        "required": ["contract_no", "contract_name", "client_id", "counterparty", "start_date", "end_date"],
        "unique": ["contract_no"],
    },
    "固定资产": {
        "model": FixedAsset,
        "columns": ["name", "category", "purchase_date", "original_value",
                     "residual_value", "useful_life_months", "monthly_depreciation",
                     "accumulated_depreciation", "net_value", "status", "client_id", "remark"],
        "required": ["name", "category", "purchase_date", "original_value", "client_id"],
        "unique": [],
    },
}

HEADER_LABELS = {
    "name": "公司名称", "tax_no": "统一社会信用代码", "taxpayer_type": "纳税人类型(general/small)",
    "industry": "行业分类", "address": "地址", "contact_person": "联系人",
    "contact_phone": "联系电话", "service_start": "服务起始日", "service_end": "合同到期日",
    "remark": "备注", "is_active": "是否启用(True/False)",
    "code": "科目编码", "category": "科目类别(资产/负债/权益/收入/费用)",
    "parent_code": "上级科目编码", "direction": "借贷方向(debit/credit)",
    "voucher_no": "凭证号", "voucher_date": "凭证日期", "summary": "摘要",
    "entries": "分录(JSON)", "total_debit": "借方合计", "total_credit": "贷方合计",
    "status": "状态", "client_id": "客户ID", "created_by": "创建者",
    "buyer_name": "购方名称", "buyer_tax_no": "购方税号", "buyer_address": "购方地址",
    "buyer_phone": "购方电话", "buyer_bank": "购方开户行", "buyer_account": "购方账号",
    "invoice_type": "发票类型", "items": "商品明细(JSON)", "total_amount": "不含税金额",
    "total_tax": "税额", "grand_total": "价税合计", "invoice_code": "发票代码",
    "invoice_no": "发票号码",
    "tax_type": "税种", "period": "所属期", "filing_result": "申报结果(JSON)", "filed_at": "申报日期",
    "contract_no": "合同编号", "contract_name": "合同名称", "contract_type": "合同类型",
    "counterparty": "对方单位", "amount": "合同金额", "start_date": "开始日期",
    "end_date": "结束日期", "payment_terms": "付款条款", "revenue_period": "收入确认期间",
    "monthly_revenue": "月均确认收入",
    "purchase_date": "购置日期", "original_value": "原值",
    "residual_value": "残值", "useful_life_months": "使用年限(月)",
    "monthly_depreciation": "月折旧额", "accumulated_depreciation": "累计折旧",
    "net_value": "净值",
}

# ============================================================
# 导出
# ============================================================

EXPORT_SHEETS = ["客户信息", "会计科目", "记账凭证", "发票", "申报记录", "合同", "固定资产"]


def export_all(db: Session) -> io.BytesIO:
    """导出全部业务数据为 Excel 工作簿"""
    wb = Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for sheet_name in EXPORT_SHEETS:
        ws = wb.create_sheet(title=sheet_name)
        defn = SHEET_DEFS[sheet_name]
        model = defn["model"]
        columns = defn["columns"]
        rows = db.query(model).all()

        # 表头
        for col_idx, col in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=HEADER_LABELS.get(col, col))
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # 数据
        for row_idx, obj in enumerate(rows, 2):
            for col_idx, col in enumerate(columns, 1):
                val = getattr(obj, col, None)
                if isinstance(val, (datetime, date)):
                    val = val.isoformat()
                elif isinstance(val, bool):
                    val = str(val)
                elif val is None:
                    val = ""
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = thin_border

        # 列宽
        for col_idx in range(1, len(columns) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 18

        ws.auto_filter.ref = ws.dimensions

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ============================================================
# 导入
# ============================================================

def parse_excel(file_bytes: bytes) -> dict[str, list[dict]]:
    """解析 Excel 文件，返回 {sheet_name: [row_dict, ...]}"""
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    result = {}
    for sheet_name in EXPORT_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            result[sheet_name] = []
            continue
        # 第一行是中文表头，映射回字段名
        headers = [str(h).strip() if h else "" for h in rows[0]]
        label_to_field = {v: k for k, v in HEADER_LABELS.items()}
        field_map = []
        for h in headers:
            field_map.append(label_to_field.get(h, None))

        data = []
        for row in rows[1:]:
            if all(v is None or str(v).strip() == "" for v in row):
                continue
            record = {}
            for idx, val in enumerate(row):
                if idx >= len(field_map):
                    break
                field = field_map[idx]
                if field is None:
                    continue
                if val is None:
                    val = None
                else:
                    val = str(val).strip()
                    if val == "":
                        val = None
                record[field] = val
            data.append(record)
        result[sheet_name] = data
    return result


def preview_import(file_bytes: bytes) -> dict:
    """预览导入数据，返回各 sheet 的行数、验证结果"""
    parsed = parse_excel(file_bytes)
    preview = {}
    for sheet_name, rows in parsed.items():
        defn = SHEET_DEFS.get(sheet_name)
        if not defn:
            continue
        issues = []
        for i, row in enumerate(rows):
            # 检查必填字段
            for req in defn["required"]:
                if not row.get(req):
                    issues.append(f"第{i+2}行: 缺少必填字段「{HEADER_LABELS.get(req, req)}」")
        preview[sheet_name] = {
            "count": len(rows),
            "columns": defn["columns"],
            "issues": issues,
            "sample": rows[:3] if rows else [],
        }
    total = sum(p["count"] for p in preview.values())
    return {"sheets": preview, "total_rows": total}


def execute_import(db: Session, file_bytes: bytes, conflict_strategy: str = "skip") -> dict:
    """执行数据导入

    conflict_strategy:
      - "skip": 遇到重复跳过（默认）
      - "update": 遇到重复更新
      - "overwrite": 清空该sheet后全量导入
    """
    parsed = parse_excel(file_bytes)
    report = {"imported": {}, "skipped": {}, "errors": {}, "total_imported": 0, "total_skipped": 0}

    for sheet_name in EXPORT_SHEETS:
        defn = SHEET_DEFS.get(sheet_name)
        rows = parsed.get(sheet_name, [])
        if not rows:
            continue

        model = defn["model"]
        unique_cols = defn.get("unique", [])
        imported = 0
        skipped = 0
        errors = []

        if conflict_strategy == "overwrite":
            db.query(model).delete()
            db.flush()

        for i, row in enumerate(rows):
            try:
                # 转换类型
                converted = _convert_row(row, defn)

                # 检查唯一性冲突
                existing = None
                if unique_cols and conflict_strategy != "overwrite":
                    for uc in unique_cols:
                        if converted.get(uc):
                            existing = db.query(model).filter(
                                getattr(model, uc) == converted[uc]
                            ).first()
                            if existing:
                                break

                if existing:
                    if conflict_strategy == "skip":
                        skipped += 1
                        continue
                    elif conflict_strategy == "update":
                        for k, v in converted.items():
                            if v is not None or k not in ["id"]:
                                setattr(existing, k, v)
                        db.flush()
                        imported += 1
                        continue

                # 生成 ID
                if "id" not in converted or not converted["id"]:
                    converted["id"] = uuid.uuid4().hex

                obj = model(**converted)
                db.add(obj)
                db.flush()
                imported += 1
            except Exception as e:
                errors.append(f"第{i+2}行: {str(e)[:120]}")

        report["imported"][sheet_name] = imported
        report["skipped"][sheet_name] = skipped
        report["errors"][sheet_name] = errors
        report["total_imported"] += imported
        report["total_skipped"] += skipped

    db.commit()
    return report


def _convert_row(row: dict, defn: dict) -> dict:
    """将字符串行数据转换为模型字段类型"""
    converted = {}
    for col in defn["columns"]:
        val = row.get(col)
        if val is None:
            converted[col] = None
            continue

        # 日期字段
        if col in ("voucher_date", "service_start", "service_end", "filed_at",
                    "start_date", "end_date", "purchase_date"):
            try:
                converted[col] = date.fromisoformat(val) if val else None
            except (ValueError, TypeError):
                converted[col] = None
            continue

        # 布尔字段
        if col in ("is_active",):
            converted[col] = val.lower() in ("true", "1", "yes") if isinstance(val, str) else bool(val)
            continue

        # 数值字段
        if col in ("total_debit", "total_credit", "amount", "monthly_revenue",
                    "original_value", "residual_value", "monthly_depreciation",
                    "accumulated_depreciation", "net_value", "total_amount", "total_tax",
                    "grand_total", "useful_life_months"):
            try:
                converted[col] = float(val) if val else 0.0
            except (ValueError, TypeError):
                converted[col] = 0.0
            continue

        converted[col] = val
    return converted


def generate_template() -> io.BytesIO:
    """生成导入模板（含表头和示例数据）"""
    wb = Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    example_fill = PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid")

    for sheet_name in EXPORT_SHEETS:
        ws = wb.create_sheet(title=sheet_name)
        defn = SHEET_DEFS[sheet_name]
        columns = defn["columns"]

        # 表头
        for col_idx, col in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=HEADER_LABELS.get(col, col))
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # 示例数据（第一条为灰色示例）
        example = _get_example_row(sheet_name)
        if example:
            for col_idx, col in enumerate(columns, 1):
                val = example.get(col, "")
                cell = ws.cell(row=2, column=col_idx, value=val)
                cell.fill = example_fill
                cell.font = Font(color="6B7280", italic=True)

        # 列宽
        for col_idx in range(1, len(columns) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 20

        ws.auto_filter.ref = ws.dimensions

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _get_example_row(sheet_name: str) -> dict | None:
    examples = {
        "客户信息": {"name": "示例科技有限公司", "tax_no": "91110108MA01ABCDE7",
                  "taxpayer_type": "small", "industry": "信息技术",
                  "address": "北京市朝阳区示例路1号", "contact_person": "张三",
                  "contact_phone": "13800138000", "is_active": "True"},
        "会计科目": {"code": "1001", "name": "库存现金", "category": "资产",
                  "direction": "debit", "is_active": "True"},
        "记账凭证": {"voucher_no": "JZ-2026-0001", "voucher_date": "2026-06-01",
                  "summary": "计提本月工资", "entries": '[{"account_code":"5001","account_name":"管理费用","debit":10000,"credit":0}]',
                  "total_debit": "10000", "total_credit": "10000", "status": "draft", "created_by": "manual"},
        "发票": {"client_id": "c001", "buyer_name": "示例科技有限公司",
                "buyer_tax_no": "91110108MA01ABCDE7", "invoice_type": "electronic_normal",
                "items": '[{"name":"技术服务","amount":10000,"tax_rate":0.06}]',
                "total_amount": "10000", "total_tax": "600", "grand_total": "10600", "status": "draft"},
        "申报记录": {"tax_type": "vat", "period": "2026-05", "client_id": "c001",
                  "status": "pending"},
        "合同": {"client_id": "c001", "contract_no": "HT-2026-001",
                "contract_name": "代理记账服务合同", "contract_type": "service",
                "counterparty": "示例科技有限公司", "amount": "12000",
                "start_date": "2026-01-01", "end_date": "2026-12-31", "status": "active"},
        "固定资产": {"name": "办公电脑", "category": "电子设备", "purchase_date": "2026-01-15",
                  "original_value": "8000", "residual_value": "400",
                  "useful_life_months": "36", "status": "in_use", "client_id": "c001"},
    }
    return examples.get(sheet_name)


# ============================================================
# 完整数据库导出（JSON格式，用于完整迁移）
# ============================================================

def export_full_dump(db: Session) -> dict:
    """导出全部数据为结构化 JSON，包含关联关系"""
    clients = db.query(Client).all()
    accounts = db.query(ChartOfAccount).all()
    vouchers = db.query(AccountingVoucher).all()
    invoices = db.query(Invoice).all()
    filings = db.query(TaxFiling).all()
    contracts = db.query(Contract).all()
    assets = db.query(FixedAsset).all()

    def _row(obj, fields):
        d = {}
        for f in fields:
            v = getattr(obj, f, None)
            if isinstance(v, (datetime, date)):
                v = v.isoformat()
            d[f] = v
        return d

    return {
        "version": "2.1.0",
        "exported_at": datetime.now().isoformat(),
        "clients": [_row(c, ["id","name","tax_no","taxpayer_type","industry","address",
                              "contact_person","contact_phone","service_start","service_end",
                              "remark","is_active"]) for c in clients],
        "accounts": [_row(a, ["id","code","name","category","parent_code","direction","is_active"])
                     for a in accounts],
        "vouchers": [_row(v, ["id","voucher_no","voucher_date","summary","entries",
                               "total_debit","total_credit","status","client_id","created_by"])
                     for v in vouchers],
        "invoices": [_row(iv, ["id","client_id","buyer_name","buyer_tax_no","invoice_type",
                                "items","total_amount","total_tax","grand_total","status",
                                "invoice_code","invoice_no","created_by"])
                     for iv in invoices],
        "filings": [_row(f, ["id","tax_type","period","client_id","filing_result","status","filed_at"])
                    for f in filings],
        "contracts": [_row(c, ["id","client_id","contract_no","contract_name","contract_type",
                                "counterparty","amount","start_date","end_date","status"])
                      for c in contracts],
        "fixed_assets": [_row(a, ["id","name","category","purchase_date","original_value",
                                   "residual_value","useful_life_months","monthly_depreciation",
                                   "accumulated_depreciation","net_value","status","client_id"])
                         for a in assets],
        "counts": {
            "clients": len(clients), "accounts": len(accounts),
            "vouchers": len(vouchers), "invoices": len(invoices),
            "filings": len(filings), "contracts": len(contracts),
            "fixed_assets": len(assets),
        }
    }


def import_full_dump(db: Session, data: dict, conflict_strategy: str = "skip") -> dict:
    """从完整 JSON dump 导入数据"""
    mappings = [
        ("clients", Client, ["id","name","tax_no","taxpayer_type","industry","address",
                              "contact_person","contact_phone","service_start","service_end",
                              "remark","is_active"], ["tax_no"]),
        ("accounts", ChartOfAccount, ["id","code","name","category","parent_code",
                                       "direction","is_active"], ["code"]),
        ("vouchers", AccountingVoucher, ["id","voucher_no","voucher_date","summary","entries",
                                          "total_debit","total_credit","status","client_id",
                                          "created_by"], ["voucher_no"]),
        ("invoices", Invoice, ["id","client_id","buyer_name","buyer_tax_no","invoice_type",
                                "items","total_amount","total_tax","grand_total","status",
                                "invoice_code","invoice_no","created_by"], []),
        ("filings", TaxFiling, ["id","tax_type","period","client_id","filing_result",
                                 "status","filed_at"], []),
        ("contracts", Contract, ["id","client_id","contract_no","contract_name","contract_type",
                                  "counterparty","amount","start_date","end_date","status"],
         ["contract_no"]),
        ("fixed_assets", FixedAsset, ["id","name","category","purchase_date","original_value",
                                       "residual_value","useful_life_months","monthly_depreciation",
                                       "accumulated_depreciation","net_value","status","client_id"],
         []),
    ]

    report = {"imported": {}, "skipped": {}, "total_imported": 0, "total_skipped": 0}

    for key, model, fields, uniques in mappings:
        rows = data.get(key, [])
        imported = 0
        skipped = 0

        if conflict_strategy == "overwrite" and rows:
            db.query(model).delete()

        for row in rows:
            converted = {}
            for f in fields:
                v = row.get(f)
                if v and f in ("voucher_date","service_start","service_end","filed_at",
                               "start_date","end_date","purchase_date"):
                    try:
                        converted[f] = date.fromisoformat(v)
                    except (ValueError, TypeError):
                        converted[f] = v
                else:
                    converted[f] = v

            # 检查唯一冲突
            existing = None
            if conflict_strategy != "overwrite":
                for uc in uniques:
                    if converted.get(uc):
                        existing = db.query(model).filter(
                            getattr(model, uc) == converted[uc]
                        ).first()
                        if existing:
                            break

            if existing:
                if conflict_strategy == "skip":
                    skipped += 1
                    continue
                elif conflict_strategy == "update":
                    for k, v in converted.items():
                        if v is not None:
                            setattr(existing, k, v)
                    imported += 1
                    continue

            if "id" not in converted or not converted.get("id"):
                converted["id"] = uuid.uuid4().hex

            db.add(model(**converted))
            imported += 1

        report["imported"][key] = imported
        report["skipped"][key] = skipped
        report["total_imported"] += imported
        report["total_skipped"] += skipped

    db.commit()
    return report


def validate_dump(data: dict) -> list[str]:
    """校验 JSON dump 结构完整性"""
    issues = []
    required_keys = ["clients", "accounts", "vouchers", "invoices", "filings", "contracts", "fixed_assets"]
    for k in required_keys:
        if k not in data:
            issues.append(f"缺少数据块: {k}")
        elif not isinstance(data[k], list):
            issues.append(f"数据块 {k} 应为数组")
    if "version" not in data:
        issues.append("缺少版本号字段 version")
    return issues


# ============================================================
# 亿企赢 Excel 格式适配
# ============================================================

# 亿企赢导出的常见列名 → 本系统字段名
YQY_COLUMN_MAP = {
    # 客户信息
    "客户名称": "name", "公司名称": "name", "企业名称": "name", "纳税人名称": "name",
    "统一社会信用代码": "tax_no", "纳税人识别号": "tax_no", "税号": "tax_no",
    "纳税人类型": "taxpayer_type", "纳税人性质": "taxpayer_type",
    "行业": "industry", "所属行业": "industry", "行业分类": "industry",
    "注册地址": "address", "经营地址": "address", "地址": "address",
    "联系人": "contact_person", "联系人姓名": "contact_person",
    "联系电话": "contact_phone", "手机号": "contact_phone", "电话": "contact_phone",
    "服务开始日期": "service_start", "服务起始": "service_start",
    "合同到期日": "service_end", "服务截止": "service_end", "到期日期": "service_end",
    "备注": "remark", "说明": "remark",
    "状态": "is_active", "启用": "is_active",

    # 会计科目
    "科目编码": "code", "科目代码": "code", "编码": "code",
    "科目名称": "name", "名称": "name",
    "科目类别": "category", "类别": "category",
    "上级科目": "parent_code", "上级编码": "parent_code",
    "借贷方向": "direction", "方向": "direction",
    "是否启用": "is_active",

    # 记账凭证
    "凭证号": "voucher_no", "凭证编号": "voucher_no",
    "凭证日期": "voucher_date", "日期": "voucher_date",
    "摘要": "summary", "凭证摘要": "summary",
    "分录": "entries", "分录明细": "entries", "凭证分录": "entries",
    "借方金额": None, "借方合计": "total_debit", "借方": "total_debit",
    "贷方金额": None, "贷方合计": "total_credit", "贷方": "total_credit",
    "凭证状态": "status", "审核状态": "status",
    "客户ID": "client_id", "所属客户": "client_id",
    "制单人": "created_by", "创建人": "created_by",

    # 发票
    "购方名称": "buyer_name", "购买方名称": "buyer_name",
    "购方税号": "buyer_tax_no", "购买方税号": "buyer_tax_no",
    "购方地址": "buyer_address", "购方电话": "buyer_phone",
    "购方开户行": "buyer_bank", "购方账号": "buyer_account",
    "发票类型": "invoice_type", "票据类型": "invoice_type",
    "商品明细": "items", "货物明细": "items",
    "不含税金额": "total_amount", "金额": "total_amount",
    "税额": "total_tax", "税额合计": "total_tax",
    "价税合计": "grand_total", "含税金额": "grand_total",
    "发票代码": "invoice_code", "发票号码": "invoice_no",
    "开票状态": "status", "发票状态": "status",

    # 申报记录
    "税种": "tax_type", "申报税种": "tax_type",
    "所属期": "period", "申报所属期": "period", "税款所属期": "period",
    "申报结果": "filing_result", "申报数据": "filing_result",
    "申报状态": "status",
    "申报日期": "filed_at", "申报时间": "filed_at",
    "客户": "client_id",

    # 合同
    "合同编号": "contract_no", "合同号": "contract_no",
    "合同名称": "contract_name", "合同标题": "contract_name",
    "合同类型": "contract_type",
    "对方单位": "counterparty", "对方": "counterparty", "签约方": "counterparty",
    "合同金额": "amount", "金额": "amount",
    "开始日期": "start_date", "起始日期": "start_date",
    "结束日期": "end_date", "截止日期": "end_date",
    "合同状态": "status",
    "付款方式": "payment_terms", "付款条款": "payment_terms",
    "收入确认": "revenue_period",

    # 固定资产
    "资产名称": "name", "名称": "name",
    "资产类别": "category", "分类": "category",
    "购置日期": "purchase_date", "购买日期": "purchase_date",
    "原值": "original_value", "原价": "original_value",
    "残值": "residual_value", "残值率": None,
    "使用年限": "useful_life_months", "折旧年限": "useful_life_months",
    "月折旧额": "monthly_depreciation",
    "累计折旧": "accumulated_depreciation",
    "净值": "net_value", "账面净值": "net_value",
    "资产状态": "status",
    "所属客户": "client_id",
}

# 根据列头识别属于哪个 Sheet
YQY_SHEET_SIGNATURES = {
    "客户信息": ["客户名称", "公司名称", "纳税人识别号", "统一社会信用代码"],
    "会计科目": ["科目编码", "科目代码", "科目名称", "科目类别"],
    "记账凭证": ["凭证号", "凭证编号", "凭证日期", "摘要"],
    "发票": ["发票代码", "发票号码", "购方名称", "发票类型"],
    "申报记录": ["税种", "申报税种", "所属期", "申报所属期", "税款所属期"],
    "合同": ["合同编号", "合同名称", "对方单位", "签约方"],
    "固定资产": ["资产名称", "资产类别", "购置日期", "原值", "固定资产"],
}


def detect_and_parse_yqy(file_bytes: bytes) -> dict:
    """检测亿企赢 Excel 格式 → 映射为标准字段 → 返回 parsed dict"""
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    result = {}

    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue

        # 第一行为列头（中文）
        raw_headers = [str(h).strip() if h else "" for h in rows[0]]

        # 尝试识别此 sheet 对应哪个类型
        sheet_type = _detect_sheet_type(raw_headers)

        # 将中文列头映射为本系统字段名
        field_map = []
        unmapped = []
        for h in raw_headers:
            field = YQY_COLUMN_MAP.get(h)
            if field:
                field_map.append(field)
            else:
                # 模糊匹配
                matched = _fuzzy_match_column(h)
                if matched:
                    field_map.append(matched)
                else:
                    field_map.append(None)
                    if h:
                        unmapped.append(h)

        # 提取数据行
        data = []
        for row in rows[1:]:
            if all(v is None or str(v).strip() == "" for v in row):
                continue
            record = {}
            for idx, val in enumerate(row):
                if idx >= len(field_map):
                    break
                field = field_map[idx]
                if field is None:
                    continue
                if val is not None:
                    val = str(val).strip()
                    if val == "":
                        val = None
                record[field] = val
            if record:
                data.append(record)

        sheet_name = sheet_type or ws.title
        # 如果已存在同类型 sheet，追加数据
        if sheet_name in result:
            result[sheet_name].extend(data)
        else:
            result[sheet_name] = data

        if unmapped:
            log_error("YQY", f"Sheet 「{ws.title}」未映射列: {unmapped[:10]}")

    return result


def _detect_sheet_type(headers: list[str]) -> str | None:
    """根据列头识别 sheet 类型"""
    for sheet_type, signatures in YQY_SHEET_SIGNATURES.items():
        for sig in signatures:
            if sig in headers:
                return sheet_type
    # 模糊匹配
    header_str = " ".join(headers)
    if any(kw in header_str for kw in ["客户", "公司", "纳税人", "企业"]):
        return "客户信息"
    if any(kw in header_str for kw in ["科目", "编码"]):
        return "会计科目"
    if any(kw in header_str for kw in ["凭证", "分录", "借方", "贷方"]):
        return "记账凭证"
    if any(kw in header_str for kw in ["发票", "票据", "购方"]):
        return "发票"
    if any(kw in header_str for kw in ["申报", "税种", "所属期"]):
        return "申报记录"
    if any(kw in header_str for kw in ["合同", "签约", "对方"]):
        return "合同"
    if any(kw in header_str for kw in ["资产", "折旧", "原值"]):
        return "固定资产"
    return None


def _fuzzy_match_column(header: str) -> str | None:
    """模糊匹配中文列名 → 字段名"""
    if not header:
        return None
    # 尝试部分匹配
    for cn_name, field in YQY_COLUMN_MAP.items():
        if field and (cn_name in header or header in cn_name):
            return field
    return None


def import_yqy_files(db: Session, files: list, conflict_strategy: str = "skip") -> dict:
    """
    批量导入亿企赢采集的文件
    files: list of YQYFile (category + content bytes)
    """
    total_report = {"imported": {}, "skipped": {}, "errors": {}, "total_imported": 0, "total_skipped": 0}

    for f in files:
        try:
            # 先尝试亿企赢格式解析
            parsed = detect_and_parse_yqy(f.content)
            if not parsed or all(len(v) == 0 for v in parsed.values()):
                # 回退到标准格式解析
                parsed = parse_excel(f.content)

            for sheet_name, rows in parsed.items():
                if not rows:
                    continue
                defn = SHEET_DEFS.get(sheet_name)
                if not defn:
                    continue

                model = defn["model"]
                unique_cols = defn.get("unique", [])
                imported = 0
                skipped = 0

                for row in rows:
                    try:
                        converted = _convert_row(row, defn)
                        existing = None
                        if unique_cols and conflict_strategy != "overwrite":
                            for uc in unique_cols:
                                if converted.get(uc):
                                    existing = db.query(model).filter(
                                        getattr(model, uc) == converted[uc]
                                    ).first()
                                    if existing:
                                        break

                        if existing:
                            if conflict_strategy == "skip":
                                skipped += 1
                                continue
                            elif conflict_strategy == "update":
                                for k, v in converted.items():
                                    if v is not None:
                                        setattr(existing, k, v)
                                imported += 1
                                continue

                        if "id" not in converted or not converted.get("id"):
                            converted["id"] = uuid.uuid4().hex

                        db.add(model(**converted))
                        imported += 1
                    except Exception as e:
                        total_report["errors"].setdefault(sheet_name, []).append(str(e)[:120])

                total_report["imported"][sheet_name] = total_report["imported"].get(sheet_name, 0) + imported
                total_report["skipped"][sheet_name] = total_report["skipped"].get(sheet_name, 0) + skipped
                total_report["total_imported"] += imported
                total_report["total_skipped"] += skipped

        except Exception as e:
            total_report["errors"].setdefault(f.category, []).append(str(e)[:200])

    db.commit()
    return total_report
