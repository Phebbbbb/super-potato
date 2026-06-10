"""
亿企赢/亿企代账 自动数据采集器
基于 Playwright，自动登录 → 导出 → 下载 Excel → 解析

用法:
  scraper = YQYScraper(headless=True)
  result = await scraper.run(username="xxx", password="xxx", org_name="xxx公司")
  # result.files 包含下载的 Excel 文件路径
"""
import asyncio
import io
import os
import re
import tempfile
import uuid
from datetime import datetime
from pathlib import Path
from typing import Optional
from dataclasses import dataclass, field

from app.services.playwright_helpers import (
    detect_captcha, retry_with_backoff, safe_goto, save_debug_snapshot,
)
from app.services.error_handler import log_error, log_info


# ============================================================
# 亿企赢平台配置
# ============================================================

YQY_CONFIG = {
    "name": "亿企赢·亿企代账",
    "login_url": "https://app.17win.com/login",
    # 备用登录入口
    "alt_login_urls": [
        "https://daizhang.17win.com/login",
        "https://app.17win.com",
        "https://www.17win.com/login",
    ],
    "login_selectors": {
        # 用户名/手机号输入框
        "username": (
            'input[name="username"], input[name="mobile"], input[name="phone"], '
            'input[placeholder*="手机"], input[placeholder*="账号"], input[placeholder*="用户名"], '
            'input[type="text"][class*="login"], input[class*="username"]'
        ),
        # 密码输入框
        "password": (
            'input[name="password"], input[type="password"], '
            'input[placeholder*="密码"], input[class*="password"]'
        ),
        # 登录按钮
        "submit": (
            'button:has-text("登录"), button:has-text("登 录"), button[type="submit"], '
            'a:has-text("登录"), div[class*="login-btn"], button[class*="login"]'
        ),
        # 可能的企业选择/切换入口
        "org_selector": (
            'select[name="org"], .org-selector, .company-select, '
            'div[class*="org"], div[class*="company"], '
            'input[placeholder*="企业"], input[placeholder*="公司"]'
        ),
    },
    # 数据导出页面导航
    "nav_selectors": {
        # 客户管理
        "client_list": (
            'text=客户管理, text=客户列表, a:has-text("客户"), '
            'span:has-text("客户"), div:has-text("客户管理")'
        ),
        "client_export": (
            'text=导出, button:has-text("导出"), button:has-text("下载"), '
            'a:has-text("导出"), span:has-text("导出")'
        ),
        # 记账 / 凭证
        "voucher_menu": (
            'text=记账, text=凭证管理, text=凭证查询, '
            'a:has-text("凭证"), a:has-text("记账"), span:has-text("凭证")'
        ),
        "voucher_export": (
            'text=导出凭证, button:has-text("导出"), button:has-text("下载"), '
            'a:has-text("导出"), span:has-text("导出")'
        ),
        # 税务申报
        "filing_menu": (
            'text=纳税申报, text=申报管理, text=申报查询, '
            'a:has-text("申报"), span:has-text("申报")'
        ),
        "filing_export": (
            'text=导出申报, button:has-text("导出"), button:has-text("下载"), '
            'a:has-text("导出")'
        ),
        # 发票
        "invoice_menu": (
            'text=发票管理, text=发票台账, text=开票管理, '
            'a:has-text("发票"), span:has-text("发票")'
        ),
        "invoice_export": (
            'text=导出, button:has-text("导出"), button:has-text("下载"), '
            'a:has-text("导出发票")'
        ),
        # 科目/账套
        "account_menu": (
            'text=科目管理, text=会计科目, text=账套设置, '
            'a:has-text("科目"), span:has-text("科目")'
        ),
        "account_export": (
            'text=导出, button:has-text("导出"), button:has-text("下载"), '
            'a:has-text("导出")'
        ),
        # 合同
        "contract_menu": (
            'text=合同管理, text=合同列表, '
            'a:has-text("合同"), span:has-text("合同")'
        ),
        "contract_export": (
            'text=导出, button:has-text("导出"), button:has-text("下载"), '
            'a:has-text("导出")'
        ),
    },
    # 下载等待超时
    "download_timeout_ms": 60000,
    # 页面加载等待
    "page_load_timeout_ms": 30000,
    # 操作间隔（避免反爬）
    "action_delay_ms": 800,
}


# ============================================================
# 采集结果
# ============================================================

@dataclass
class YQYFile:
    """下载的文件"""
    category: str       # clients / accounts / vouchers / invoices / filings / contracts
    filename: str
    content: bytes
    row_count: int = 0


@dataclass
class YQYResult:
    """亿企赢采集结果"""
    success: bool
    files: list = field(default_factory=list)    # list[YQYFile]
    errors: list = field(default_factory=list)
    screenshots: list = field(default_factory=list)
    duration_seconds: float = 0
    summary: str = ""


class YQYScraper:
    """亿企赢 Playwright 采集器"""

    def __init__(self, headless: bool = True, screenshot_dir: str = None):
        self.headless = headless
        self.screenshot_dir = screenshot_dir or os.path.join(
            os.path.dirname(__file__), "..", "..", "screenshots", "yqy"
        )
        os.makedirs(self.screenshot_dir, exist_ok=True)
        self._browser = None
        self._context = None
        self._page = None

    async def run(
        self,
        username: str,
        password: str,
        org_name: str = None,
        categories: list = None,
    ) -> YQYResult:
        """一键运行：登录 → 遍历导出 → 下载文件"""
        start = datetime.now()
        result = YQYResult(success=False)
        if categories is None:
            categories = ["clients", "accounts", "vouchers", "invoices", "filings", "contracts"]

        try:
            from playwright.async_api import async_playwright

            async with async_playwright() as p:
                browser = await p.chromium.launch(headless=self.headless)
                context = await browser.new_context(
                    accept_downloads=True,
                    viewport={"width": 1400, "height": 900},
                )
                page = await context.new_page()
                self._browser = browser
                self._context = context
                self._page = page

                # Step 1: 登录
                ss = await self._screenshot(page, "01_login_start")
                result.screenshots.append(ss)

                logged_in = await self._login(page, username, password)
                if not logged_in:
                    result.errors.append("登录失败：请检查账号密码是否正确")
                    await browser.close()
                    return result

                ss = await self._screenshot(page, "02_logged_in")
                result.screenshots.append(ss)
                log_info("YQY", "登录成功")

                # Step 2: 选择企业（如有多个）
                if org_name:
                    await self._select_org(page, org_name)

                # Step 3: 遍历导出各模块
                for cat in categories:
                    try:
                        file_data = await self._export_category(page, cat, context)
                        if file_data:
                            result.files.append(file_data)
                            log_info("YQY", f"导出 {cat} — {file_data.row_count} 行")
                        else:
                            log_error("YQY", f"导出 {cat} — 无数据")
                    except Exception as e:
                        msg = f"导出 {cat} 失败: {str(e)[:120]}"
                        result.errors.append(msg)
                        log_error("YQY", msg)
                        ss = await self._screenshot(page, f"error_{cat}")
                        result.screenshots.append(ss)

                await browser.close()

        except ImportError:
            result.errors.append("Playwright 未安装，请执行: pip install playwright && playwright install chromium")
        except Exception as e:
            result.errors.append(f"采集异常: {str(e)[:200]}")

        result.success = len(result.files) > 0
        result.duration_seconds = (datetime.now() - start).total_seconds()
        result.summary = (
            f"采集完成：{len(result.files)} 个数据模块, "
            f"{sum(f.row_count for f in result.files)} 条记录, "
            f"耗时 {result.duration_seconds:.0f} 秒"
        )
        return result

    # ----- 登录 -----

    async def _login(self, page, username: str, password: str) -> bool:
        """尝试多个已知登录 URL，返回是否成功"""
        urls = [YQY_CONFIG["login_url"]] + YQY_CONFIG["alt_login_urls"]

        for url in urls:
            try:
                await safe_goto(page, url, timeout=YQY_CONFIG["page_load_timeout_ms"])
                await asyncio.sleep(0.5)
            except Exception:
                continue

            selectors = YQY_CONFIG["login_selectors"]

            # 填充用户名
            filled_user = await self._fill_field(page, selectors["username"], username)
            if not filled_user:
                continue  # 可能不是登录页，试下一个 URL

            # 填充密码
            await self._fill_field(page, selectors["password"], password)
            await asyncio.sleep(0.3)

            # 检测验证码
            has_captcha = await detect_captcha(page)
            if has_captcha:
                await save_debug_snapshot(page, "yqy_captcha_detected")
                log_error("YQY", "检测到验证码，需要人工处理")
                return False

            # 点击登录
            clicked = await self._click_first(page, selectors["submit"])
            if not clicked:
                continue

            # 等待登录完成
            await asyncio.sleep(2)
            try:
                await page.wait_for_load_state("networkidle", timeout=10000)
            except Exception:
                pass

            # 检查是否登录成功（页面上有客户/记账/申报等菜单）
            success_indicators = [
                'text=客户管理', 'text=记账', 'text=申报',
                'text=工作台', 'text=首页',
                '.main-menu', '.nav', '.sidebar',
            ]
            for indicator in success_indicators:
                try:
                    el = await page.wait_for_selector(indicator, timeout=3000)
                    if el:
                        return True
                except Exception:
                    continue

            # 如果页面上还有登录框，说明失败了
            try:
                still_login = await page.query_selector('input[type="password"]')
                if still_login:
                    continue
            except Exception:
                pass

            return True

        return False

    # ----- 企业选择 -----

    async def _select_org(self, page, org_name: str):
        """在多企业账号中选择指定企业"""
        selectors = YQY_CONFIG["login_selectors"]["org_selector"]
        try:
            # 尝试下拉选择
            for sel in selectors.split(","):
                sel = sel.strip()
                try:
                    option = await page.wait_for_selector(
                        f'{sel} option:has-text("{org_name}")', timeout=3000
                    )
                    if option:
                        value = await option.get_attribute("value")
                        if value:
                            await page.select_option(sel, value)
                            await asyncio.sleep(0.5)
                            return
                except Exception:
                    pass

            # 尝试点击文本
            await self._click_first(page, f'text={org_name}, a:has-text("{org_name}")')
            await asyncio.sleep(0.5)
        except Exception:
            pass

    # ----- 分类导出 -----

    async def _export_category(self, page, category: str, context) -> Optional[YQYFile]:
        """导航到指定分类 → 触发导出 → 等待下载"""
        nav = YQY_CONFIG["nav_selectors"]

        menu_sel = nav.get(f"{category}_menu", "")
        export_sel = nav.get(f"{category}_export", "")

        if not menu_sel or not export_sel:
            return None

        # 1. 点击菜单
        clicked = await self._click_first(page, menu_sel)
        if not clicked:
            return None
        await asyncio.sleep(YQY_CONFIG["action_delay_ms"] / 1000)
        await self._wait_stable(page)

        # 2. 点击导出按钮
        download_promise = None
        try:
            # 设置下载监听
            async with page.expect_download(timeout=YQY_CONFIG["download_timeout_ms"]) as download_info:
                await self._click_first(page, export_sel)
                await asyncio.sleep(1)  # 等待导出弹窗/确认

                # 可能还有确认按钮
                await self._click_first(page, 'text=确认导出, text=确定, button:has-text("确定"), button:has-text("导出")')
                await asyncio.sleep(0.5)

            download = await download_info.value
            filename = download.suggested_filename or f"{category}.xlsx"
            content = await download.read()

            # 解析 Excel 行数
            row_count = self._count_excel_rows(content)

            return YQYFile(
                category=category,
                filename=filename,
                content=content,
                row_count=row_count,
            )
        except Exception as e:
            # 下载超时或未触发，尝试点击后手动查找新出现的下载链接
            log_error("YQY", f"下载监听失败: {str(e)[:80]}")
            return None

    # ----- 辅助方法 -----

    async def _fill_field(self, page, selector_str: str, value: str) -> bool:
        """尝试多个选择器填充输入框"""
        for sel in selector_str.split(","):
            sel = sel.strip()
            try:
                el = await page.wait_for_selector(sel, timeout=3000)
                if el:
                    await el.click()
                    await el.fill("")
                    await el.type(value, delay=50)
                    return True
            except Exception:
                continue
        return False

    async def _click_first(self, page, selector_str: str) -> bool:
        """尝试多个选择器点击第一个匹配元素"""
        for sel in selector_str.split(","):
            sel = sel.strip()
            try:
                el = await page.wait_for_selector(sel, timeout=3000)
                if el:
                    await el.click()
                    await asyncio.sleep(0.3)
                    return True
            except Exception:
                continue
        return False

    async def _wait_stable(self, page):
        """等待页面稳定"""
        try:
            await page.wait_for_load_state("networkidle", timeout=5000)
        except Exception:
            pass
        await asyncio.sleep(0.5)

    async def _screenshot(self, page, name: str) -> str:
        """截图并返回路径"""
        ts = datetime.now().strftime("%H%M%S")
        path = os.path.join(self.screenshot_dir, f"yqy_{ts}_{name}.png")
        try:
            await page.screenshot(path=path, full_page=False)
        except Exception:
            pass
        return path

    def _count_excel_rows(self, content: bytes) -> int:
        """快速统计 Excel 行数"""
        try:
            import openpyxl
            from io import BytesIO
            wb = openpyxl.load_workbook(BytesIO(content), read_only=True)
            total = sum(ws.max_row - 1 for ws in wb.worksheets if ws.max_row)  # 减去表头
            wb.close()
            return max(0, total)
        except Exception:
            return 0


# ============================================================
# 同步包装器（供 API 调用）
# ============================================================

async def scrape_yqy(
    username: str,
    password: str,
    org_name: str = None,
    categories: list = None,
    headless: bool = True,
) -> YQYResult:
    scraper = YQYScraper(headless=headless)
    return await scraper.run(username, password, org_name, categories)


def scrape_yqy_sync(
    username: str,
    password: str,
    org_name: str = None,
    categories: list = None,
    headless: bool = True,
) -> YQYResult:
    """同步包装，方便在非 async 上下文中调用"""
    return asyncio.run(scrape_yqy(username, password, org_name, categories, headless))
